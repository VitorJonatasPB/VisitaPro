"""
Utilitários para geração de dados de relatórios.
Contém lógica reutilizável para diferentes tipos de relatórios.
"""

from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Q, Sum, Avg, Min, Max
from django.db.models.functions import TruncDate
from django.core.files.base import ContentFile
from decimal import Decimal
import json
import csv
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .models import Visita, Empresa, CustomUser, Jornada, Funcionario, VisitaFoto, RespostaRelatorio
from .relatorios_models import RelatorioGerado, ConfiguracaoRelatorio


class GeradorRelatorios:
    """Classe principal para geração de relatórios"""
    
    def __init__(self, usuario=None, data_inicio=None, data_fim=None):
        self.usuario = usuario
        self.data_inicio = data_inicio or (timezone.now() - timedelta(days=30)).date()
        self.data_fim = data_fim or timezone.now().date()
        self.periodo_dias = (self.data_fim - self.data_inicio).days + 1
    
    def gerar_resumo_geral(self):
        """Gera resumo geral de todas as métricas importantes"""
        
        visitas = Visita.objects.filter(
            data__range=[self.data_inicio, self.data_fim]
        )
        
        visitas_realizadas = visitas.filter(status='realizada').count()
        visitas_agendadas = visitas.filter(status='agendada').count()
        visitas_canceladas = visitas.filter(status='cancelada').count()
        
        empresas = Empresa.objects.all()
        empresas_ativas = empresas.filter(status='A').count()
        empresas_inativas = empresas.filter(status='I').count()
        empresas_negociacao = empresas.filter(status='N').count()
        
        # Taxa de conversão
        empresas_convertidas = empresas.filter(
            status='A',
            data_conversao__isnull=False,
            data_conversao__range=[self.data_inicio, self.data_fim]
        ).count()
        taxa_conversao = (
            (empresas_convertidas / max(empresas_negociacao, 1)) * 100
            if empresas_negociacao > 0 else 0
        )
        
        media_visitas = (
            visitas.count() / max(empresas.count(), 1)
            if empresas.count() > 0 else 0
        )
        
        assessores = CustomUser.objects.filter(is_assessor=True).count()
        
        jornadas = Jornada.objects.filter(
            data__range=[self.data_inicio, self.data_fim],
            status='finalizada'
        )
        distancia_media = jornadas.aggregate(
            avg_km=Avg('km_total')
        )['avg_km'] or 0.0
        
        return {
            'total_visitas': visitas.count(),
            'visitas_realizadas': visitas_realizadas,
            'visitas_agendadas': visitas_agendadas,
            'visitas_canceladas': visitas_canceladas,
            'total_empresas': empresas.count(),
            'empresas_ativas': empresas_ativas,
            'empresas_inativas': empresas_inativas,
            'empresas_negociacao': empresas_negociacao,
            'taxa_conversao_percentual': round(taxa_conversao, 2),
            'media_visitas_por_empresa': round(float(media_visitas), 2),
            'total_assessores': assessores,
            'distancia_media_percorrida': round(float(distancia_media), 2),
            'periodo': {
                'data_inicio': self.data_inicio.isoformat(),
                'data_fim': self.data_fim.isoformat(),
                'dias': self.periodo_dias
            }
        }
    
    def gerar_performance_assessor(self, assessor_id=None):
        """Gera dados de performance de um ou vários assessores"""
        
        assessores = CustomUser.objects.filter(is_assessor=True)
        if assessor_id:
            assessores = assessores.filter(id=assessor_id)
        
        dados = []
        
        for assessor in assessores:
            visitas = Visita.objects.filter(
                assessor=assessor,
                data__range=[self.data_inicio, self.data_fim]
            )
            
            visitas_realizadas = visitas.filter(status='realizada').count()
            taxa_realizacao = (
                (visitas_realizadas / max(visitas.count(), 1)) * 100
                if visitas.count() > 0 else 0
            )
            
            empresas_visitadas = visitas.values('empresa').distinct().count()
            empresas_ativas_assessor = visitas.filter(
                empresa__status='A'
            ).values('empresa').distinct().count()
            
            jornadas = Jornada.objects.filter(
                assessor=assessor,
                data__range=[self.data_inicio, self.data_fim],
                status='finalizada'
            )
            
            distancia_total = jornadas.aggregate(
                total=Sum('km_total')
            )['total'] or 0.0
            
            jornadas_trabalhadas = jornadas.count()
            media_visitas_dia = (
                visitas_realizadas / max(jornadas_trabalhadas, 1)
                if jornadas_trabalhadas > 0 else 0
            )
            
            ultima_atividade = visitas.aggregate(
                ultima=Max('atualizado_em')
            )['ultima'] or None
            
            dados.append({
                'assessor_id': assessor.id,
                'assessor_nome': assessor.get_full_name(),
                'total_visitas': visitas.count(),
                'visitas_realizadas': visitas_realizadas,
                'taxa_realizacao_percentual': round(taxa_realizacao, 2),
                'empresas_visitadas': empresas_visitadas,
                'empresas_ativas_assessor': empresas_ativas_assessor,
                'distancia_total_km': round(float(distancia_total), 2),
                'jornadas_trabalhadas': jornadas_trabalhadas,
                'media_visitas_dia': round(float(media_visitas_dia), 2),
                'ultima_atividade': ultima_atividade.isoformat() if ultima_atividade else None
            })
        
        return dados
    
    def gerar_status_empresas(self, empresa_id=None):
        """Gera status detalhado de empresas"""
        
        empresas = Empresa.objects.all()
        if empresa_id:
            empresas = empresas.filter(id=empresa_id)
        
        dados = []
        
        for empresa in empresas:
            visitas = empresa.visitas.filter(
                data__range=[self.data_inicio, self.data_fim]
            )
            
            dados.append({
                'empresa_id': empresa.id,
                'empresa_nome': empresa.nome,
                'status': empresa.get_status_display(),
                'assessor_responsavel': (
                    empresa.assessor.get_full_name()
                    if empresa.assessor else None
                ),
                'total_visitas': visitas.count(),
                'ultima_visita': empresa.ultima_visita.isoformat() if empresa.ultima_visita else None,
                'funcionarios_total': empresa.funcionarios.count(),
                'data_conversao': (
                    empresa.data_conversao.isoformat()
                    if empresa.data_conversao else None
                ),
                'localizacao': {
                    'latitude': empresa.latitude or 'N/A',
                    'longitude': empresa.longitude or 'N/A',
                    'cidade': empresa.cidade or 'N/A',
                    'estado': empresa.estado or 'N/A'
                }
            })
        
        return dados
    
    def gerar_visitas_detalhadas(self, empresa_id=None, assessor_id=None):
        """Gera dados detalhados das visitas"""
        
        visitas = Visita.objects.filter(
            data__range=[self.data_inicio, self.data_fim]
        ).select_related('empresa', 'assessor')
        
        if empresa_id:
            visitas = visitas.filter(empresa_id=empresa_id)
        if assessor_id:
            visitas = visitas.filter(assessor_id=assessor_id)
        
        dados = []
        
        for visita in visitas:
            respostas_formulario = []
            for resposta in visita.respostas.all():
                respostas_formulario.append({
                    'pergunta': resposta.pergunta.texto,
                    'resposta': resposta.resposta
                })
            
            # Calcular duração
            duracao = None
            if visita.checkin_time and visita.checkout_time:
                duracao = int(
                    (visita.checkout_time - visita.checkin_time).total_seconds() / 60
                )
            
            dados.append({
                'visita_id': visita.id,
                'empresa_nome': visita.empresa.nome,
                'assessor_nome': visita.assessor.get_full_name(),
                'data': visita.data.isoformat(),
                'horario': visita.horario.isoformat() if visita.horario else None,
                'status': visita.get_status_display(),
                'duracao_minutes': duracao,
                'funcionarios_atendidos': visita.contatoes_atendidos.count(),
                'fotos_quantidade': visita.fotos.count(),
                'respostas_formulario': respostas_formulario,
                'observacoes': visita.observacoes or '',
                'criado_em': visita.criado_em.isoformat()
            })
        
        return dados
    
    def gerar_jornadas_resumo(self, assessor_id=None):
        """Gera resumo de jornadas e quilometragem"""
        
        jornadas = Jornada.objects.filter(
            data__range=[self.data_inicio, self.data_fim],
            status='finalizada'
        ).select_related('assessor')
        
        if assessor_id:
            jornadas = jornadas.filter(assessor_id=assessor_id)
        
        dados = []
        
        for jornada in jornadas:
            visitas_dia = Visita.objects.filter(
                assessor=jornada.assessor,
                data=jornada.data,
                status='realizada'
            ).count()
            
            dados.append({
                'jornada_id': jornada.id,
                'assessor_nome': jornada.assessor.get_full_name(),
                'data': jornada.data.isoformat(),
                'hora_inicio': jornada.inicio_time.time().isoformat() if jornada.inicio_time else None,
                'hora_fim': jornada.fim_time.time().isoformat() if jornada.fim_time else None,
                'status': jornada.get_status_display(),
                'quilometragem_total': jornada.km_total,
                'visitas_realizadas': visitas_dia,
                'localizacao_inicio': {
                    'latitude': jornada.inicio_lat or 'N/A',
                    'longitude': jornada.inicio_lng or 'N/A'
                },
                'localizacao_fim': {
                    'latitude': jornada.fim_lat or 'N/A',
                    'longitude': jornada.fim_lng or 'N/A'
                } if jornada.fim_lat else None
            })
        
        return dados
    
    def gerar_conversoes(self):
        """Gera análise de conversão de empresas de negociação para ativa"""
        
        empresas_convertidas = Empresa.objects.filter(
            data_conversao__isnull=False,
            data_conversao__range=[self.data_inicio, self.data_fim]
        )
        
        dados = []
        
        for empresa in empresas_convertidas:
            primeira_visita = empresa.visitas.aggregate(
                primeira=Min('data')
            )['primeira']
            
            visitas_pre = empresa.visitas.filter(
                data__lt=empresa.data_conversao
            ).count()
            
            visitas_pos = empresa.visitas.filter(
                data__gte=empresa.data_conversao
            ).count()
            
            dias_conversao = None
            if primeira_visita and empresa.data_conversao:
                dias_conversao = (empresa.data_conversao - primeira_visita).days
            
            dados.append({
                'empresa_id': empresa.id,
                'empresa_nome': empresa.nome,
                'data_primeiro_contato': primeira_visita.isoformat() if primeira_visita else None,
                'data_conversao': empresa.data_conversao.isoformat(),
                'dias_para_conversao': dias_conversao,
                'visitas_pre_conversao': visitas_pre,
                'visitas_pos_conversao': visitas_pos,
                'status_atual': empresa.get_status_display(),
                'assessor_responsavel': (
                    empresa.assessor.get_full_name()
                    if empresa.assessor else 'Sem assessor'
                )
            })
        
        return dados
    
    def gerar_feedback_visitas(self):
        """Gera relatório de feedback e respostas das visitas"""
        
        visitas = Visita.objects.filter(
            data__range=[self.data_inicio, self.data_fim],
            status='realizada'
        ).select_related('empresa')
        
        dados = []
        
        for visita in visitas:
            respostas = []
            for resposta in visita.respostas.all():
                respostas.append({
                    'pergunta': resposta.pergunta.texto,
                    'resposta': resposta.resposta,
                    'tipo': resposta.pergunta.tipo_resposta
                })
            
            # Verificar se formulário está completo
            total_perguntas_ativas = RespostaRelatorio.objects.filter(
                visita=visita
            ).count()
            
            dados.append({
                'visita_id': visita.id,
                'empresa_nome': visita.empresa.nome,
                'data_visita': visita.data.isoformat(),
                'respostas': respostas,
                'observacoes': visita.observacoes or '',
                'formulario_completo': len(respostas) > 0,
                'total_respostas': len(respostas)
            })
        
        return dados


class ExportadorRelatorios:
    """Classe para exportar relatórios em diferentes formatos"""
    
    @staticmethod
    def exportar_json(dados, titulo="Relatório"):
        """Exporta dados em formato JSON"""
        return json.dumps(dados, indent=2, ensure_ascii=False, default=str)
    
    @staticmethod
    def exportar_csv(dados_lista, titulo="Relatório"):
        """Exporta dados em formato CSV"""
        if not dados_lista:
            return ""
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=dados_lista[0].keys())
        writer.writeheader()
        writer.writerows(dados_lista)
        
        return output.getvalue()
    
    @staticmethod
    def exportar_excel(dados_lista, titulo="Relatório"):
        """Exporta dados em formato Excel (requer openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"
            
            if not dados_lista:
                return None
            
            # Cabeçalhos
            headers = list(dados_lista[0].keys())
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Dados
            for row_num, row_data in enumerate(dados_lista, 2):
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = row_data.get(header, "")
            
            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:
            return None
    
    @staticmethod
    def exportar_pdf(dados, titulo="Relatório", nome_arquivo="relatorio.pdf"):
        """Exporta dados em formato PDF"""
        output = io.BytesIO()
        
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Título
        titulo_para = Paragraph(titulo, title_style)
        elements.append(titulo_para)
        elements.append(Spacer(1, 0.3*inch))
        
        # Verificar tipo de dados
        if isinstance(dados, dict):
            # Dados estruturados (resumo, performance, etc)
            tabela_dados = []
            for chave, valor in dados.items():
                if isinstance(valor, dict):
                    continue  # Pular dicionários aninhados
                tabela_dados.append([str(chave).replace('_', ' ').title(), str(valor)])
            
            if tabela_dados:
                tabela = Table(tabela_dados, colWidths=[3.5*inch, 2.5*inch])
                tabela.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                elements.append(tabela)
        
        elif isinstance(dados, list) and dados:
            # Lista de registros
            headers = list(dados[0].keys())
            tabela_dados = [headers]
            
            for item in dados[:10]:  # Limitar a 10 linhas na visualização
                row = [str(item.get(h, ""))[:30] for h in headers]
                tabela_dados.append(row)
            
            tabela = Table(tabela_dados)
            tabela.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(tabela)
        
        doc.build(elements)
        output.seek(0)
        return output.getvalue()
